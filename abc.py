import telebot
import random
import json
import os
from datetime import datetime
from telebot import types
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    import gspread
except ImportError:
    gspread = None

# --- НАСТРОЙКИ ---
BOT_TOKEN = os.getenv("BOT_TOKEN", "7981486621:AAGewMxW4yi_N6BQTbUypJnsDcARM1XyLNE")
# Список администраторов
ADMIN_IDS = [1655296176, 5735687272, 8344814435]

BTN_MY_ORDERS = "Мои заказы"
BTN_CONTACT_ADMIN = "Связь с админом"
BTN_CREATE_ORDER = "Создать заказ"
BTN_ADMIN_PANEL = "Админ-панель"
BTN_BACK = "Назад"
BTN_BACK_TO_MENU = "↩️ В админ-меню"
BTN_MY_ACCEPTED = "📦 Сегодняшние заказы"

BUTTON_TEXTS = [BTN_MY_ORDERS, BTN_CONTACT_ADMIN, BTN_CREATE_ORDER, BTN_ADMIN_PANEL, "/start"]
EXCEL_FILE = "log_orders.xlsx"
ACTION_LOG_FILE = "bot_actions_log.txt"
USERS_FILE = "users.json"
ORDERS_FILE = "orders.json"
TOTAL_SHELVES = 1000
GOOGLE_SERVICE_ACCOUNT_FILE = os.getenv(
    "GOOGLE_SERVICE_ACCOUNT_FILE",
    r"data-excel.json"
)
GOOGLE_SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID", "1IxSW29IAPtgSUyyNigALrMtMO1jiQoADxt0EmsnzNUY")
GOOGLE_WORKSHEET_TITLE = "Выданные заказы"
LOG_HEADERS = ["Дата", "Время", "Код заказа", "ФИО", "Телефон", "ID клиента", "Кол-во посылок", "Сумма", "10%"]
DAILY_CONFIRMED_WORKSHEET_TITLE = "Подтвержденные заказы"
DAILY_CONFIRMED_HEADERS = [
    "Дата подтверждения",
    "Время подтверждения",
    "Код заказа",
    "ФИО",
    "Телефон",
    "Кол-во посылок",
    "Telegram ID",
    "Username",
    "Статус",
    "Подтвердил админ",
    "QR file_id"
]
DAILY_CONFIRMED_DATE_CELL = f"{get_column_letter(len(DAILY_CONFIRMED_HEADERS) + 1)}1"
STATUS_PENDING_APPROVAL = "pending_admin_approval"
STATUS_ACCEPTED_BY_ADMIN = "accepted_by_admin"
STATUS_IN_TRANSIT = "in_transit"
STATUS_READY = "ready"
STATUS_ISSUED = "выдан"
CUSTOM_MESSAGE_MENU_COMMANDS = set(BUTTON_TEXTS + ["/start", "/admin"])
ORDER_STATUS_PRIORITY = {
    STATUS_READY: 0,
    STATUS_IN_TRANSIT: 1,
    STATUS_ACCEPTED_BY_ADMIN: 2,
    STATUS_PENDING_APPROVAL: 2,
    STATUS_ISSUED: 3,
}
temp_orders = {}
google_orders_worksheet = None
google_orders_error = None
google_daily_orders_worksheet = None
google_daily_orders_error = None
bot = telebot.TeleBot(BOT_TOKEN)


# --- РАБОТА С ДАННЫМИ ---
def load_data(file):
    if not os.path.exists(file):
        with open(file, 'w', encoding='utf-8') as f:
            json.dump({}, f)
    with open(file, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
            if file == ORDERS_FILE:
                data, changed = normalize_orders_data(data)
                data, expired_orders = expire_outdated_daily_orders(data)
                if changed or expired_orders:
                    save_data(file, data)
                notify_about_expired_daily_orders(expired_orders)
            return data
        except:
            return {}


def save_data(file, data):
    with open(file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def normalize_orders_data(data):
    if not isinstance(data, dict):
        return {}, True

    changed = False
    for order_data in data.values():
        if not isinstance(order_data, dict):
            continue

        if order_data.get("status") == STATUS_PENDING_APPROVAL:
            order_data["status"] = STATUS_ACCEPTED_BY_ADMIN
            if not order_data.get("accepted_at"):
                order_data["accepted_at"] = order_data.get("created_at") or datetime.now().isoformat()
            changed = True

        if "accepted_by_admin" not in order_data:
            order_data["accepted_by_admin"] = None
            changed = True

        if "accepted_at" not in order_data:
            order_data["accepted_at"] = None
            changed = True

        if "admin_review_messages" not in order_data:
            order_data["admin_review_messages"] = {}
            changed = True

    return data, changed


def parse_order_datetime(value):
    if not value:
        return None

    try:
        return datetime.fromisoformat(value)
    except:
        return None


def expire_outdated_daily_orders(data):
    current_date = datetime.now().date()
    expired_orders = []

    if not isinstance(data, dict):
        return {}, expired_orders

    for code, order_data in list(data.items()):
        if not isinstance(order_data, dict):
            continue

        if order_data.get("status") not in {STATUS_ACCEPTED_BY_ADMIN, STATUS_PENDING_APPROVAL}:
            continue

        created_at = parse_order_datetime(order_data.get("created_at"))
        if created_at is None:
            continue

        if created_at.date() < current_date:
            expired_orders.append((code, dict(order_data)))
            del data[code]

    return data, expired_orders


def notify_about_expired_daily_orders(expired_orders):
    for code, order_data in expired_orders:
        try:
            bot.send_message(
                order_data["user_id"],
                f"⚠️ Заказ <code>{code}</code> удален, потому что за сегодняшний день его не забрали, а код уже недействителен.\n"
                f"Пожалуйста, создайте новый заказ.",
                parse_mode="HTML"
            )
        except:
            pass

        log_action(
            order_data.get("user_id"),
            "daily_order_expired_and_deleted",
            f"code={code}",
            username=normalize_order_value(order_data.get("username")),
            full_name=order_data.get("fio")
        )


def sanitize_log_text(value):
    if value is None:
        return "-"

    text = str(value).replace("\r", " ").replace("\n", " | ").strip()
    return text or "-"


def append_action_log_line(line):
    try:
        with open(ACTION_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(line + "\n")
    except Exception as e:
        print(f"Ошибка записи action-лога: {e}")


def log_action(user_id, action, details="", username=None, full_name=None):
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    role = "ADMIN" if is_admin(user_id) else "USER"
    username_value = sanitize_log_text(username or "нет")
    if username_value not in {"нет", "-"} and not username_value.startswith("@"):
        username_value = f"@{username_value}"
    username_text = username_value
    full_name_text = sanitize_log_text(full_name or "Нет имени")
    details_text = sanitize_log_text(details)
    append_action_log_line(
        f"[{timestamp}] {role} id={user_id} username={username_text} name={full_name_text} action={sanitize_log_text(action)} details={details_text}"
    )


def log_message_event(message, action, details=""):
    user = getattr(message, "from_user", None)
    if user is None:
        append_action_log_line(
            f"[{datetime.now().strftime('%d.%m.%Y %H:%M:%S')}] UNKNOWN action={sanitize_log_text(action)} details={sanitize_log_text(details)}"
        )
        return

    full_name = " ".join(part for part in [user.first_name, user.last_name] if part) or "Нет имени"
    log_action(user.id, action, details, username=user.username, full_name=full_name)


def log_callback_event(call, details=""):
    user = getattr(call, "from_user", None)
    action = f"callback:{call.data}"
    if user is None:
        append_action_log_line(
            f"[{datetime.now().strftime('%d.%m.%Y %H:%M:%S')}] UNKNOWN action={sanitize_log_text(action)} details={sanitize_log_text(details)}"
        )
        return

    full_name = " ".join(part for part in [user.first_name, user.last_name] if part) or "Нет имени"
    log_action(user.id, action, details, username=user.username, full_name=full_name)


def is_blocked(uid):
    users = load_data(USERS_FILE)
    return users.get(str(uid), {}).get('blocked', False)


ORDER_RETRY_TEXT = (
    "При оформлении заказа произошла ошибка. "
    "Пожалуйста, попробуйте сделать заказ снова."
)


def normalize_order_value(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    return text


def order_has_invalid_user_fields(order_data):
    return any(
        normalize_order_value(order_data.get(field)) is None
        for field in ("fio", "phone", "count")
    )


def send_retry_order_message(chat_id):
    bot.send_message(chat_id, ORDER_RETRY_TEXT)


def get_display_username(order_data):
    return normalize_order_value(order_data.get("username")) or "Нет юзернейма"


def get_order_codes_for_action(orders, code):
    order_data = orders.get(code)
    if not order_data:
        return []

    if is_admin(order_data.get("user_id")):
        return [code]

    shelf_num = order_data.get("shelf")
    if not shelf_num:
        return [code]

    return [
        current_code for current_code, data in orders.items()
        if data.get("shelf") == shelf_num and data.get("status") != STATUS_ISSUED
    ]


def is_menu_or_command(text):
    return bool(text) and text.strip() in CUSTOM_MESSAGE_MENU_COMMANDS


def build_back_keyboard(callback_data="admin_menu", label=BTN_BACK):
    return types.InlineKeyboardMarkup().add(types.InlineKeyboardButton(label, callback_data=callback_data))


def build_main_menu_keyboard(user_id):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add(BTN_MY_ORDERS, BTN_CONTACT_ADMIN)
    kb.add(BTN_CREATE_ORDER)
    if is_admin(user_id):
        kb.add(BTN_ADMIN_PANEL)
    return kb


def build_admin_panel_keyboard():
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(
        types.InlineKeyboardButton("📋 Заказы в системе", callback_data="orders_list|0"),
        types.InlineKeyboardButton("🔍 Найти заказ", callback_data="admin_info_search"),
        types.InlineKeyboardButton("👤 Найти клиента", callback_data="admin_find_by_id"),
        types.InlineKeyboardButton(BTN_MY_ACCEPTED, callback_data="accepted_nav|0"),
        types.InlineKeyboardButton("📥 Принять в ПВЗ", callback_data="admin_receive"),
        types.InlineKeyboardButton("📤 Выдать заказ", callback_data="admin_giveout_start"),
        types.InlineKeyboardButton("✉️ Написать клиенту", callback_data="admin_msg"),
        types.InlineKeyboardButton("🔒 Доступ и блокировка", callback_data="admin_access")
    )
    return kb


def make_order_record(user_id, order_code, draft):
    return {
        "user_id": user_id,
        "status": STATUS_ACCEPTED_BY_ADMIN,
        "file_id": draft["file_id"],
        "fio": draft["fio"],
        "phone": draft["phone"],
        "count": draft["count"],
        "shelf": None,
        "username": draft["username"],
        "created_at": datetime.now().isoformat(),
        "accepted_by_admin": None,
        "accepted_at": None,
        "admin_review_messages": {},
        "order_code": order_code
    }


def mark_order_as_approved(order_data, admin_id):
    order_data["status"] = STATUS_ACCEPTED_BY_ADMIN
    order_data["accepted_by_admin"] = admin_id
    order_data["accepted_at"] = datetime.now().isoformat()


def mark_order_as_in_transit(order_data, admin_id=None):
    order_data["status"] = STATUS_IN_TRANSIT
    if admin_id is not None:
        order_data["accepted_by_admin"] = admin_id
        order_data["accepted_at"] = datetime.now().isoformat()


def mark_order_as_ready(order_data, shelf_num, price):
    order_data["shelf"] = shelf_num
    order_data["status"] = STATUS_READY
    order_data["price"] = price


def mark_order_as_issued(order_data):
    order_data["status"] = STATUS_ISSUED
    order_data["shelf"] = None


def is_admin(user_id):
    return user_id in ADMIN_IDS


def delete_message_safe(chat_id, message_id):
    try:
        bot.delete_message(chat_id, message_id)
    except:
        pass


def delete_messages_safe(chat_id, *message_ids):
    for message_id in message_ids:
        if message_id:
            delete_message_safe(chat_id, message_id)


def cleanup_transient_input(message, *prompt_message_ids):
    delete_messages_safe(message.chat.id, *prompt_message_ids, message.message_id)


def cleanup_prompt_messages(chat_id, *prompt_message_ids):
    delete_messages_safe(chat_id, *prompt_message_ids)


def replace_message_content(chat_id, message_id, text, reply_markup=None, parse_mode="HTML"):
    try:
        bot.edit_message_caption(
            text,
            chat_id=chat_id,
            message_id=message_id,
            parse_mode=parse_mode,
            reply_markup=reply_markup
        )
        return True
    except:
        try:
            bot.edit_message_text(
                text,
                chat_id=chat_id,
                message_id=message_id,
                parse_mode=parse_mode,
                reply_markup=reply_markup
            )
            return True
        except:
            return False


def get_order_price_value(order_data):
    raw_price = normalize_order_value(order_data.get("price"))
    if raw_price is None:
        return 0.0

    try:
        return float(raw_price.replace(",", "."))
    except:
        return 0.0


def has_order_price(order_data):
    return normalize_order_value(order_data.get("price")) is not None


def place_order_on_shelf(chat_id, code, orders, back_kb):
    user_id = orders[code]['user_id']
    existing_shelf = None
    if not is_admin(user_id):
        for current_code, order_data in orders.items():
            if current_code == code:
                continue
            if order_data.get('user_id') == user_id and order_data.get('shelf') and order_data.get('status') != STATUS_ISSUED:
                existing_shelf = order_data['shelf']
                break

    free_shelf = existing_shelf if existing_shelf else get_free_shelf()
    if existing_shelf:
        bot.send_message(
            chat_id,
            f"ℹ️ У этого клиента уже есть заказ в ячейке № {existing_shelf}. Кладите туда же!"
        )

    if not free_shelf:
        bot.send_message(
            chat_id,
            f"⛔️ Склад переполнен! Нет свободных полок ({TOTAL_SHELVES}/{TOTAL_SHELVES}).",
            reply_markup=back_kb
        )
        return

    raw_price = orders[code].get("price")
    mark_order_as_ready(orders[code], free_shelf, raw_price)
    log_action(
        user_id,
        "order_placed_on_shelf",
        f"code={code}; shelf={free_shelf}; price={sanitize_log_text(raw_price)}",
        username=normalize_order_value(orders[code].get("username")),
        full_name=orders[code].get("fio")
    )
    save_data(ORDERS_FILE, orders)

    commission = int(get_order_price_value(orders[code]) // 10)
    try:
        bot.send_message(
            orders[code]['user_id'],
            f"🎉 Ваш заказ {code} прибыл на ПВЗ! К оплате {commission} р.\nМожно приходить забирать."
        )
    except:
        pass

    msg = bot.send_message(
        chat_id,
        f"✅ Заказ <code>{code}</code> размещен в ячейке № <b>{free_shelf}</b>.\nЖду следующий код:",
        parse_mode="HTML",
        reply_markup=back_kb
    )
    bot.register_next_step_handler(msg, process_receive_mass, msg.message_id)


def get_order_status_text(order_data, for_user=False):
    status = order_data.get("status")

    if status == STATUS_PENDING_APPROVAL:
        return "📦 Ожидает в ПВЗ в Донецке"
    if status == STATUS_ACCEPTED_BY_ADMIN:
        return "📦 Ожидает в ПВЗ в Донецке" if not for_user else "📦 Заказ оформлен и ожидает в ПВЗ в Донецке"
    if status == STATUS_IN_TRANSIT:
        return "🚚 Едет в ПВЗ"
    if status == STATUS_ISSUED:
        return "✅ Выдан"
    if order_data.get("shelf"):
        return "📍 Готов к выдаче" if for_user else f"🚪 Ячейка: {order_data['shelf']}"
    return "⏳ Ожидает поступления"


def get_order_status_info_text(order_data):
    status = order_data.get("status")

    if status == STATUS_PENDING_APPROVAL:
        return "📦 Ожидает в ПВЗ в Донецке"
    if status == STATUS_ACCEPTED_BY_ADMIN:
        return "📦 Заказ оформлен и ожидает в ПВЗ в Донецке"
    if status == STATUS_IN_TRANSIT:
        admin_id = order_data.get("accepted_by_admin")
        if admin_id:
            return f"🚚 Заказ забрал администратор <code>{admin_id}</code>, он едет в ПВЗ"
        return "🚚 Заказ забрали, он едет в ПВЗ"
    if status == STATUS_ISSUED:
        return "✅ Заказ уже выдан"
    if order_data.get("shelf"):
        return f"🚪 <b>Ячейка: {order_data['shelf']}</b>"
    return "⏳ Еще не принят (без полки)"


def build_order_card_lines(order_code, order_data, status_text=None):
    lines = [
        f"🔢 Код: <code>{order_code}</code>",
        f"👤 ФИО: {order_data.get('fio', '')}",
        f"📞 Телефон: <code>{order_data.get('phone', '')}</code>",
        f"📦 Кол-во: {order_data.get('count', '')}",
        f"🆔 TG ID: <code>{order_data.get('user_id', '')}</code>",
        f"🔗 Username: {get_display_username(order_data)}",
    ]

    status_text = status_text or get_order_status_text(order_data)
    if status_text:
        lines.append(f"📌 Статус: {status_text}")

    return lines


def build_order_detail_caption(title, order_code, order_data, status_text=None, extra_lines=None):
    lines = [title, ""]
    lines.extend(build_order_card_lines(order_code, order_data, status_text=status_text))

    if extra_lines:
        prepared_extra_lines = [line for line in extra_lines if line]
        if prepared_extra_lines:
            lines.extend([""] + prepared_extra_lines)

    return "\n".join(lines)


def build_user_orders_text(user_id, orders):
    user_codes = get_sorted_order_codes(orders, filter_ids=user_id)
    if not user_codes:
        return "У вас пока нет заказов."

    lines = ["📦 <b>Ваши заказы:</b>", ""]
    for code in user_codes:
        order_data = orders[code]
        lines.append(f"🔢 Код: <code>{code}</code>")
        lines.append(get_order_status_text(order_data, for_user=True))
        lines.append(f"👤 {order_data.get('fio', '')} | 📞 <code>{order_data.get('phone', '')}</code>")
        lines.append("")

    return "\n".join(lines).strip()


def build_admin_order_info_keyboard(order_code):
    kb = types.InlineKeyboardMarkup(row_width=2)
    kb.add(
        types.InlineKeyboardButton("❌ Удалить", callback_data=f"del_confirm|{order_code}"),
        types.InlineKeyboardButton(BTN_BACK, callback_data="admin_menu")
    )
    kb.add(
        types.InlineKeyboardButton("✉️ Связаться с клиентом", callback_data=f"msg_client|{order_code}"),
        types.InlineKeyboardButton("👤 Инфо о клиенте", callback_data=f"client_info|{order_code}")
    )
    return kb


def build_admin_review_caption(order_code, order_data, footer_text=None):
    return build_order_detail_caption(
        "🆕 <b>Новый заказ</b>",
        order_code,
        order_data,
        extra_lines=[footer_text]
    )


def build_taken_order_caption(order_code, order_data, index, total):
    return build_order_detail_caption(
        f"📦 <b>Сегодняшние задачи {index + 1}/{total}</b>",
        order_code,
        order_data
    )


def get_taken_order_codes(orders, admin_id):
    codes = [
        code for code, data in orders.items()
        if data.get("status") == STATUS_ACCEPTED_BY_ADMIN
    ]
    return sorted(codes, key=lambda code: orders[code].get("created_at") or orders[code].get("accepted_at") or "")


def build_taken_orders_keyboard(order_code, index, total):
    kb = types.InlineKeyboardMarkup(row_width=2)

    nav_buttons = []
    if index > 0:
        nav_buttons.append(types.InlineKeyboardButton("⬅️ Назад", callback_data=f"accepted_nav|{index - 1}"))
    if index < total - 1:
        nav_buttons.append(types.InlineKeyboardButton("➡️ Вперед", callback_data=f"accepted_nav|{index + 1}"))
    if nav_buttons:
        kb.row(*nav_buttons)

    kb.add(types.InlineKeyboardButton("🚚 Принять заказ", callback_data=f"accepted_pickup|{order_code}|{index}"))
    kb.add(types.InlineKeyboardButton("✉️ Написать пользователю", callback_data=f"accepted_message|{order_code}|{index}"))
    kb.add(types.InlineKeyboardButton("❌ Удалить заказ", callback_data=f"accepted_delete|{order_code}|{index}"))
    kb.add(types.InlineKeyboardButton("↩️ В админ-меню", callback_data="admin_menu"))
    return kb


def edit_or_send_admin_photo_message(chat_id, order_data, caption, reply_markup=None, message_id=None):
    if message_id is not None:
        media = types.InputMediaPhoto(order_data["file_id"], caption=caption, parse_mode="HTML")
        try:
            bot.edit_message_media(media=media, chat_id=chat_id, message_id=message_id, reply_markup=reply_markup)
            return message_id
        except:
            try:
                bot.edit_message_caption(caption=caption, chat_id=chat_id, message_id=message_id,
                                         parse_mode="HTML", reply_markup=reply_markup)
                return message_id
            except:
                pass

    try:
        sent_message = bot.send_photo(chat_id, order_data["file_id"], caption=caption,
                                      parse_mode="HTML", reply_markup=reply_markup)
        return sent_message.message_id
    except:
        return None


def sync_admin_review_messages(order_code, order_data, footer_text, reply_markup=None):
    caption = build_admin_review_caption(order_code, order_data, footer_text)
    stored_ids = order_data.get("admin_review_messages", {})
    updated_ids = {}

    for admin_id in ADMIN_IDS:
        message_id = edit_or_send_admin_photo_message(
            admin_id,
            order_data,
            caption,
            reply_markup=reply_markup,
            message_id=stored_ids.get(str(admin_id))
        )
        if message_id:
            updated_ids[str(admin_id)] = message_id

    return updated_ids


def send_order_for_admin_review(order_code, order_data):
    return sync_admin_review_messages(
        order_code,
        order_data,
        "📦 Новый заказ от пользователя. Он уже добавлен в общую очередь заказов ."
    )


def build_daily_confirmed_row(order_code, order_data, admin_id):
    now = datetime.now()
    return [
        now.strftime("%d.%m.%Y"),
        now.strftime("%H:%M:%S"),
        order_code,
        order_data.get("fio", ""),
        order_data.get("phone", ""),
        order_data.get("count", ""),
        order_data.get("user_id", ""),
        get_display_username(order_data),
        order_data.get("status", ""),
        admin_id,
        order_data.get("file_id", "")
    ]


def get_unique_code():
    """Генерирует уникальный 4-значный код заказа."""
    orders = load_data(ORDERS_FILE)
    while True:
        code = "".join([str(random.randint(0, 9)) for _ in range(4)])
        if code not in orders:
            return code


def get_free_shelf():
    """Ищет первую свободную полку от 1 до 80."""
    orders = load_data(ORDERS_FILE)
    # Собираем все занятые полки
    occupied_shelves = {data['shelf'] for data in orders.values() if data.get('shelf')}

    for i in range(1, TOTAL_SHELVES + 1):
        if i not in occupied_shelves:
            return i
    return None  # Если всё занято


def get_sorted_order_codes(orders_dict, filter_ids=None):
    """
    Сортирует заказы по этапам жизненного цикла:
    1. Уже на полке (по номеру полки)
    2. В пути в ПВЗ
    3. В очереди на забор у клиента
    4. Выданные
    """

    def sort_key(code):
        d = orders_dict[code]
        status = d.get("status")
        shelf = d.get("shelf")
        created_at = d.get("created_at") or ""
        accepted_at = d.get("accepted_at") or created_at

        if shelf:
            return (ORDER_STATUS_PRIORITY[STATUS_READY], int(shelf), created_at, code)

        priority = ORDER_STATUS_PRIORITY.get(status, ORDER_STATUS_PRIORITY[STATUS_ACCEPTED_BY_ADMIN])
        timestamp = accepted_at if status == STATUS_IN_TRANSIT else created_at
        return (priority, 0, timestamp, code)

    codes = list(orders_dict.keys())
    if filter_ids:
        codes = [c for c in codes if orders_dict[c].get('user_id') == filter_ids]

    return sorted(codes, key=sort_key)

def notify_admins(text, photo_id=None):
    for admin_id in ADMIN_IDS:
        try:
            if photo_id:
                bot.send_photo(admin_id, photo_id, caption=text, parse_mode='HTML')
            else:
                bot.send_message(admin_id, text, parse_mode='HTML')
        except:
            pass


def build_log_row(order_code, order_data):
    now = datetime.now()
    date_str = now.strftime("%d.%m.%Y")
    time_str = now.strftime("%H:%M:%S")

    price = get_order_price_value(order_data)
    commission = int(price // 10)

    return [
        date_str,
        time_str,
        order_code,
        order_data.get('fio', ''),
        order_data.get('phone', ''),
        order_data.get('user_id', ''),
        order_data.get('count', 0),
        price,
        commission
    ]


def get_google_orders_worksheet():
    global google_orders_worksheet
    global google_orders_error

    if google_orders_worksheet is not None:
        return google_orders_worksheet

    if google_orders_error is not None:
        return None

    if gspread is None:
        google_orders_error = "Библиотека gspread не установлена."
        print(google_orders_error)
        return None

    if not os.path.exists(GOOGLE_SERVICE_ACCOUNT_FILE):
        google_orders_error = f"Не найден файл сервисного аккаунта: {GOOGLE_SERVICE_ACCOUNT_FILE}"
        print(google_orders_error)
        return None

    try:
        client = gspread.service_account(filename=GOOGLE_SERVICE_ACCOUNT_FILE)
        spreadsheet = client.open_by_key(GOOGLE_SPREADSHEET_ID)

        try:
            worksheet = spreadsheet.worksheet(GOOGLE_WORKSHEET_TITLE)
        except gspread.WorksheetNotFound:
            worksheets = spreadsheet.worksheets()
            if len(worksheets) == 1 and worksheets[0].title == "Лист1" and not worksheets[0].get_all_values():
                worksheet = worksheets[0]
                worksheet.update_title(GOOGLE_WORKSHEET_TITLE)
            else:
                worksheet = spreadsheet.add_worksheet(title=GOOGLE_WORKSHEET_TITLE, rows=1000, cols=len(LOG_HEADERS))

        first_row = worksheet.row_values(1)
        if first_row != LOG_HEADERS:
            if not any(first_row):
                worksheet.update([LOG_HEADERS], "A1:I1")
            else:
                worksheet.insert_row(LOG_HEADERS, 1)

        google_orders_worksheet = worksheet
        return google_orders_worksheet
    except Exception as e:
        google_orders_error = f"Ошибка подключения к Google Sheets: {e}"
        print(google_orders_error)
        return None


def get_google_daily_orders_worksheet():
    global google_daily_orders_worksheet
    global google_daily_orders_error

    if google_daily_orders_worksheet is not None:
        return google_daily_orders_worksheet

    if google_daily_orders_error is not None:
        return None

    if gspread is None:
        google_daily_orders_error = "Библиотека gspread не установлена."
        print(google_daily_orders_error)
        return None

    if not os.path.exists(GOOGLE_SERVICE_ACCOUNT_FILE):
        google_daily_orders_error = f"Не найден файл сервисного аккаунта: {GOOGLE_SERVICE_ACCOUNT_FILE}"
        print(google_daily_orders_error)
        return None

    try:
        client = gspread.service_account(filename=GOOGLE_SERVICE_ACCOUNT_FILE)
        spreadsheet = client.open_by_key(GOOGLE_SPREADSHEET_ID)

        try:
            worksheet = spreadsheet.worksheet(DAILY_CONFIRMED_WORKSHEET_TITLE)
        except gspread.WorksheetNotFound:
            worksheet = spreadsheet.add_worksheet(
                title=DAILY_CONFIRMED_WORKSHEET_TITLE,
                rows=1000,
                cols=len(DAILY_CONFIRMED_HEADERS) + 1
            )

        google_daily_orders_worksheet = worksheet
        return google_daily_orders_worksheet
    except Exception as e:
        google_daily_orders_error = f"Ошибка подключения ко второму листу Google Sheets: {e}"
        print(google_daily_orders_error)
        return None


def ensure_daily_confirmed_sheet(worksheet, current_date):
    expected_row = DAILY_CONFIRMED_HEADERS + [current_date]
    current_row = worksheet.row_values(1)
    stored_date = current_row[len(DAILY_CONFIRMED_HEADERS)] if len(current_row) > len(DAILY_CONFIRMED_HEADERS) else None
    stored_headers = current_row[:len(DAILY_CONFIRMED_HEADERS)]

    if stored_date != current_date:
        worksheet.clear()
        worksheet.update([expected_row], f"A1:{DAILY_CONFIRMED_DATE_CELL}")
        return

    if stored_headers != DAILY_CONFIRMED_HEADERS:
        worksheet.update([expected_row], f"A1:{DAILY_CONFIRMED_DATE_CELL}")


def append_row_to_google_sheet(row):
    worksheet = get_google_orders_worksheet()
    if worksheet is None:
        return

    try:
        worksheet.append_row(row, value_input_option="USER_ENTERED")
    except Exception as e:
        print(f"Ошибка записи в Google Sheets: {e}")


def append_confirmed_order_to_daily_sheet(order_code, order_data, admin_id):
    worksheet = get_google_daily_orders_worksheet()
    if worksheet is None:
        return

    try:
        today = datetime.now().strftime("%d.%m.%Y")
        ensure_daily_confirmed_sheet(worksheet, today)
        worksheet.append_row(build_daily_confirmed_row(order_code, order_data, admin_id), value_input_option="USER_ENTERED")
    except Exception as e:
        print(f"Ошибка записи подтвержденного заказа в Google Sheets: {e}")


def log_to_excel(order_code, order_data):
    row = build_log_row(order_code, order_data)
    local_error = None

    # Если файла нет, создаем его с заголовками
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(LOG_HEADERS)
    else:
        wb = None
        ws = None

    try:
        if wb is None:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
    except Exception as e:
        local_error = e

    append_row_to_google_sheet(row)

    if local_error is not None:
        raise local_error


def delete_order_and_notify_user(order_code, order_data):
    try:
        bot.send_message(
            order_data['user_id'],
            f"⚠️ Администратор удалил ваш заказ <code>{order_code}</code>. "
            f"Пожалуйста, попробуйте оформить его снова, если он еще актуален.",
            parse_mode="HTML"
        )
    except:
        pass


def show_taken_orders_view(chat_id, admin_id, index=0, message_id=None):
    orders = load_data(ORDERS_FILE)
    taken_codes = get_taken_order_codes(orders, admin_id)

    if not taken_codes:
        empty_text = "✅ Незабранных заказов не осталось."
        empty_kb = build_back_keyboard("admin_menu", BTN_BACK_TO_MENU)
        if message_id is None:
            bot.send_message(chat_id, empty_text, reply_markup=empty_kb)
            return

        if not replace_message_content(chat_id, message_id, empty_text, reply_markup=empty_kb):
            bot.send_message(chat_id, empty_text, reply_markup=empty_kb)
        return

    index = max(0, min(index, len(taken_codes) - 1))
    order_code = taken_codes[index]
    order_data = orders[order_code]
    caption = build_taken_order_caption(order_code, order_data, index, len(taken_codes))
    keyboard = build_taken_orders_keyboard(order_code, index, len(taken_codes))

    if message_id is None:
        bot.send_photo(chat_id, order_data["file_id"], caption=caption, parse_mode="HTML", reply_markup=keyboard)
        return

    edit_or_send_admin_photo_message(chat_id, order_data, caption, reply_markup=keyboard, message_id=message_id)


def send_taken_order_template(order_code, user_id, template_type):
    if template_type == "expired":
        text = (
            f"⚠️ По заказу <code>{order_code}</code> код устарел.\n"
            f"Пожалуйста, создайте заказ заново и пришлите новый код."
        )
    else:
        text = (
            f"⚠️ По заказу <code>{order_code}</code> сейчас не хватает денег на карте.\n"
            f"Пожалуйста, пополните карту и напишите через кнопку «Связь с админом», когда всё будет готово."
        )

    bot.send_message(user_id, text, parse_mode="HTML")

# --- ГЛАВНОЕ МЕНЮ КЛИЕНТА ---
def main_menu(message):
    if is_blocked(message.chat.id): return
    kb = build_main_menu_keyboard(message.chat.id)
    bot.send_message(
        message.chat.id,
        "Выберите действие ниже. После оформления заказа администратор заберет его, привезет в ПВЗ, а мы сообщим, когда можно будет забрать.",
        reply_markup=kb
    )

# --- ОФОРМЛЕНИЕ ЗАКАЗА ---
@bot.message_handler(content_types=['photo'])
def handle_qr(message):
    if is_blocked(message.chat.id): return
    log_message_event(message, "order_qr_photo_sent", "Пользователь отправил QR-фото для оформления заказа")
    current_hour = datetime.now().hour
    if (current_hour >= 16 or datetime.now().weekday() == 6) and not is_admin(message.chat.id):
        bot.send_message(message.chat.id,
                         "К сожалению, пункт выдачи заказов сейчас не принимает заказы (принимает с 0:10 до 16:00, воскресенье выходной).\n"
                         "Пожалуйста сделайте заказ в рабочее время.\n\n"
                         "❗️ ВАЖНО: когда будете оформлять заказ в следующий раз, обновите QR-код в приложении, чтобы он был сегодняшним!")
        log_message_event(message, "order_qr_rejected_by_schedule")
        return
    file_id = message.photo[-1].file_id
    msg = bot.send_message(message.chat.id, "Введите ФИО получателя:")
    bot.register_next_step_handler(msg, process_order_fio, file_id, msg.message_id)


def process_order_fio(message, file_id, prompt_message_id=None):
    if is_blocked(message.chat.id): return
    log_message_event(message, "order_fio_entered", f"fio={sanitize_log_text(message.text)}")
    cleanup_transient_input(message, prompt_message_id)
    fio = normalize_order_value(message.text)
    if fio is None:
        send_retry_order_message(message.chat.id)
        log_message_event(message, "order_fio_invalid")
        return
    msg = bot.send_message(message.chat.id, "Введите номер телефона:")
    bot.register_next_step_handler(msg, process_order_phone, file_id, fio, msg.message_id)


def process_order_phone(message, file_id, fio, prompt_message_id=None):
    if is_blocked(message.chat.id): return
    log_message_event(message, "order_phone_entered", f"phone={sanitize_log_text(message.text)}")
    cleanup_transient_input(message, prompt_message_id)
    phone = normalize_order_value(message.text)
    if phone is None:
        send_retry_order_message(message.chat.id)
        log_message_event(message, "order_phone_invalid")
        return
    msg = bot.send_message(message.chat.id, "Укажите количество посылок:")
    bot.register_next_step_handler(msg, process_order_count, file_id, fio, phone, msg.message_id)


def process_order_count(message, file_id, fio, phone, prompt_message_id=None):
    if is_blocked(message.chat.id): return
    log_message_event(message, "order_count_entered", f"count={sanitize_log_text(message.text)}")
    cleanup_transient_input(message, prompt_message_id)
    count = normalize_order_value(message.text)
    if count is None:
        send_retry_order_message(message.chat.id)
        log_message_event(message, "order_count_invalid")
        return

    order_preview = {
        "fio": fio,
        "phone": phone,
        "count": count,
    }
    if not file_id or order_has_invalid_user_fields(order_preview):
        send_retry_order_message(message.chat.id)
        log_message_event(message, "order_draft_invalid")
        return

    temp_orders[message.chat.id] = {
        "file_id": file_id,
        "fio": fio,
        "phone": phone,
        "count": count,
        "username": "@" + message.from_user.username if message.from_user.username else "Нет юзернейма"
    }
    log_message_event(
        message,
        "order_draft_created",
        f"fio={fio}; phone={phone}; count={count}"
    )

    text = (f"📝 <b>Проверьте данные вашего заказа:</b>\n\n"
            f"👤 ФИО: {fio}\n"
            f"📞 Телефон: {phone}\n"
            f"📦 Количество посылок: {count}\n\n"
            f"Всё верно?")

    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton("✅ Подтвердить", callback_data="order_confirm"),
        types.InlineKeyboardButton("❌ Отменить", callback_data="order_cancel")
    )

    bot.send_photo(message.chat.id, file_id, caption=text, parse_mode="HTML", reply_markup=kb)

    '''rand_code = get_unique_code()

    orders = load_data(ORDERS_FILE)
    orders[rand_code] = {
        "user_id": message.chat.id,
        "status": "waiting",
        "file_id": file_id,
        "fio": fio,
        "phone": phone,
        "count": count,
        "shelf": None,
        "username": "@"+str(message.from_user.username)
    }
    save_data(ORDERS_FILE, orders)

    bot.send_message(message.chat.id, f"Заказ оформлен! Ваш код получения: {rand_code}. Мы оповестим вас когда заказ придет на пункт выдачи, не удаляйте этот чат.")

    admin_info = (f"🆕 НОВЫЙ ЗАКАЗ!\n"
                  f"🔢 Код: {rand_code}\n"
                  f"👤 ФИО: {fio}\n"
                  f"📞 Тел: {phone}\n"
                  f"📦 Кол-во: {count}")
    notify_admins(admin_info, file_id)'''


# --- АДМИН-ПАНЕЛЬ ---
@bot.message_handler(commands=['admin'])
def admin_panel(message):
    if not is_admin(message.chat.id): return
    if getattr(message, "text", None) == "/admin":
        log_message_event(message, "open_admin_panel", "Команда /admin")
        delete_message_safe(message.chat.id, message.message_id)
    bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
    kb = build_admin_panel_keyboard()
    bot.send_message(message.chat.id, "Панель управления:", reply_markup=kb)


def open_admin_menu_from_callback(call):
    delete_message_safe(call.message.chat.id, call.message.message_id)
    admin_panel(call.message)


def start_receive_flow(call, back_kb):
    msg = bot.send_message(
        call.message.chat.id,
        "📥 <b>Режим приемки.</b>\nВведите код заказа:",
        parse_mode="HTML",
        reply_markup=back_kb
    )
    bot.register_next_step_handler(msg, process_receive_mass, msg.message_id)


def start_find_client_flow(call, back_kb):
    msg = bot.send_message(
        call.message.chat.id,
        "Введите Telegram ID пользователя",
        reply_markup=back_kb
    )
    bot.register_next_step_handler(msg, process_find_by_id, msg.message_id)


def cancel_contact_admin_flow(call):
    bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
    delete_message_safe(call.message.chat.id, call.message.message_id)
    bot.send_message(
        call.message.chat.id,
        "Отправка сообщения отменена.",
        reply_markup=build_main_menu_keyboard(call.message.chat.id)
    )


def start_giveout_flow(call, back_kb):
    msg = bot.send_message(
        call.message.chat.id,
        "📤 Введите код заказа:",
        reply_markup=back_kb
    )
    bot.register_next_step_handler(msg, process_giveout_search, msg.message_id)


def handle_admin_review_action(call):
    approved = call.data.startswith("admin_approve|")
    code = call.data.split("|")[1]
    orders = load_data(ORDERS_FILE)

    if code not in orders:
        replace_message_content(call.message.chat.id, call.message.message_id, "❌ Заказ уже удален.")
        return

    order_data = orders[code]
    if approved:
        if order_data.get("status") == STATUS_PENDING_APPROVAL:
            order_data["status"] = STATUS_ACCEPTED_BY_ADMIN
            save_data(ORDERS_FILE, orders)
        log_callback_event(call, f"Нажата старая кнопка подтверждения для заказа {code}")
        replace_message_content(
            call.message.chat.id,
            call.message.message_id,
            "✅ Подтверждение больше не требуется. Заказ уже находится в общей очереди заказов."
        )
        return

    log_callback_event(call, f"Удален заказ {code}")
    order_data["admin_review_messages"] = sync_admin_review_messages(code, order_data, "❌ Заказ удален.")
    delete_order_and_notify_user(code, order_data)
    del orders[code]
    save_data(ORDERS_FILE, orders)


def handle_taken_orders_action(call):
    if call.data.startswith("accepted_nav|"):
        index = int(call.data.split("|")[1])
        show_taken_orders_view(call.message.chat.id, call.from_user.id, index=index, message_id=call.message.message_id)
        return

    if call.data.startswith("accepted_pickup|"):
        _, code, index = call.data.split("|")
        index = int(index)
        orders = load_data(ORDERS_FILE)
        if code not in orders:
            show_taken_orders_view(call.message.chat.id, call.from_user.id, index=index, message_id=call.message.message_id)
            return

        order_data = orders[code]
        if order_data.get("status") != STATUS_ACCEPTED_BY_ADMIN:
            show_taken_orders_view(call.message.chat.id, call.from_user.id, index=index, message_id=call.message.message_id)
            return

        prompt = bot.send_message(
            call.message.chat.id,
            f"Введите стоимость для заказа <code>{code}</code>:",
            parse_mode="HTML"
        )
        bot.register_next_step_handler(
            prompt,
            lambda m: process_taken_order_price_input(
                m,
                code,
                index,
                prompt.message_id,
                call.message.message_id
            )
        )
        return

    if call.data.startswith("accepted_delete|"):
        _, code, index = call.data.split("|")
        index = int(index)
        orders = load_data(ORDERS_FILE)
        if code in orders and orders[code].get("status") == STATUS_ACCEPTED_BY_ADMIN:
            order_data = orders[code]
            log_callback_event(call, f"Удален принятый заказ {code}")
            delete_order_and_notify_user(code, order_data)
            del orders[code]
            save_data(ORDERS_FILE, orders)

        show_taken_orders_view(call.message.chat.id, call.from_user.id, index=index, message_id=call.message.message_id)
        return

    if call.data.startswith("accepted_message|"):
        _, code, index = call.data.split("|")
        orders = load_data(ORDERS_FILE)
        if code not in orders or orders[code].get("status") != STATUS_ACCEPTED_BY_ADMIN:
            show_taken_orders_view(call.message.chat.id, call.from_user.id, index=int(index), message_id=call.message.message_id)
            return

        kb = types.InlineKeyboardMarkup(row_width=1)
        kb.add(
            types.InlineKeyboardButton("⌛ Код устарел", callback_data=f"accepted_template|expired|{code}|{index}"),
            types.InlineKeyboardButton("💳 Нет денег на карте", callback_data=f"accepted_template|nomoney|{code}|{index}"),
            types.InlineKeyboardButton("⬅️ Назад", callback_data=f"accepted_message_back|{index}")
        )
        msg = bot.send_message(
            call.message.chat.id,
            f"Выберите шаблон для заказа <code>{code}</code> или просто отправьте свой текст пользователю.",
            parse_mode="HTML",
            reply_markup=kb
        )
        bot.register_next_step_handler(msg, lambda m: process_taken_order_message_input(m, code, int(index), msg.message_id))
        return

    if call.data.startswith("accepted_template|"):
        _, template_type, code, index = call.data.split("|")
        orders = load_data(ORDERS_FILE)
        if code not in orders:
            bot.answer_callback_query(call.id, "Заказ уже удален.")
            return

        send_taken_order_template(code, orders[code]["user_id"], template_type)
        log_callback_event(call, f"Отправлен шаблон {template_type} по заказу {code}")
        bot.answer_callback_query(call.id, "Пользователь уведомлен.")
        delete_message_safe(call.message.chat.id, call.message.message_id)
        return

    if call.data.startswith("accepted_message_back|"):
        delete_message_safe(call.message.chat.id, call.message.message_id)


def handle_issue_confirmation(call, back_kb):
    code = call.data.split("|")[1]
    orders = load_data(ORDERS_FILE)

    if code not in orders:
        bot.send_message(call.message.chat.id, "Ошибка: заказ уже удален.", reply_markup=back_kb)
        return

    if orders[code].get('status') == STATUS_ISSUED:
        bot.send_message(
            call.message.chat.id,
            f"⚠️ Заказ <code>{code}</code> уже выдан.",
            reply_markup=back_kb,
            parse_mode="HTML"
        )
        return

    if not orders[code].get('shelf'):
        bot.send_message(
            call.message.chat.id,
            f"⚠️ Заказ <code>{code}</code> еще не размещен на полке.",
            reply_markup=back_kb,
            parse_mode="HTML"
        )
        return

    shelf_num = orders[code].get('shelf', 'Нет')
    codes_on_shelf = get_order_codes_for_action(orders, code)

    total_price = 0
    for current_code in codes_on_shelf:
        total_price += get_order_price_value(orders[current_code])

    total_commission = int(total_price // 10)

    numeric_count = 0
    non_numeric_counts = []
    for current_code in codes_on_shelf:
        raw_count = str(orders[current_code].get('count') or '').strip()
        try:
            numeric_count += int(raw_count)
        except:
            non_numeric_counts.append(raw_count)

    if non_numeric_counts:
        total_count = f"{numeric_count} + {' + '.join(non_numeric_counts)}"
    else:
        total_count = str(numeric_count)

    for current_code in codes_on_shelf:
        try:
            log_to_excel(current_code, orders[current_code])
        except Exception as e:
            print(f"Ошибка записи в Excel: {e}")

        mark_order_as_issued(orders[current_code])

    log_callback_event(call, f"Выдача заказов: {', '.join(codes_on_shelf)}")
    save_data(ORDERS_FILE, orders)
    delete_message_safe(call.message.chat.id, call.message.message_id)

    issued_list = ", ".join([f"<code>{current_code}</code>" for current_code in codes_on_shelf])
    shelf_is_free = not any(
        data.get('shelf') == shelf_num and data.get('status') != STATUS_ISSUED
        for data in orders.values()
    )
    shelf_text = (
        f"🚪 Полка: {shelf_num} — свободна"
        if shelf_is_free else
        f"🚪 Полка: {shelf_num} — на ней остались другие заказы"
    )

    bot.send_message(
        call.message.chat.id,
        f"✅ Выданы заказы: {issued_list}\n"
        f"{shelf_text}\n"
        f"📦 Всего посылок: {total_count}\n"
        f"💰 К оплате (10%): {total_commission} руб.",
        reply_markup=back_kb,
        parse_mode="HTML"
    )


def show_orders_list_page(call, page):
    orders = load_data(ORDERS_FILE)
    all_codes = [code for code in get_sorted_order_codes(orders) if orders[code].get('status') != STATUS_ISSUED]
    total = len(all_codes)
    per_page = 5
    start, end = page * per_page, (page + 1) * per_page
    current_codes = all_codes[start:end]

    text = f"📋 <b>Заказы в системе ({total}):</b>\n\n"
    for code in current_codes:
        order_data = orders[code]
        text += f"🔹 <code>{code}</code> | {order_data['fio']} | <code>{order_data['phone']}</code>\n"
        text += f"   {get_order_status_text(order_data)}\n\n"

    kb = types.InlineKeyboardMarkup(row_width=2)
    if page > 0:
        kb.add(types.InlineKeyboardButton("⬅️", callback_data=f"orders_list|{page - 1}"))
    if end < total:
        kb.add(types.InlineKeyboardButton("➡️", callback_data=f"orders_list|{page + 1}"))
    kb.add(types.InlineKeyboardButton(BTN_BACK, callback_data="admin_menu"))

    if not replace_message_content(call.message.chat.id, call.message.message_id, text, reply_markup=kb):
        bot.send_message(call.message.chat.id, text, reply_markup=kb, parse_mode="HTML")


def start_order_info_search(call, back_kb):
    msg = bot.send_message(call.message.chat.id, "Введите код заказа:", reply_markup=back_kb)
    bot.register_next_step_handler(msg, process_info_search, msg.message_id)


def show_access_menu(call):
    kb = types.InlineKeyboardMarkup()
    kb.add(
        types.InlineKeyboardButton("Заблокировать", callback_data="block_user"),
        types.InlineKeyboardButton("Разблокировать", callback_data="unblock_user"),
        types.InlineKeyboardButton(BTN_BACK, callback_data="admin_menu")
    )
    replace_message_content(call.message.chat.id, call.message.message_id, "Доступ:", reply_markup=kb, parse_mode=None)


def start_block_unblock_flow(call, status, back_kb):
    prompt = "ID для бана:" if status else "ID для разбана:"
    msg = bot.send_message(call.message.chat.id, prompt, reply_markup=back_kb)
    bot.register_next_step_handler(msg, lambda m: process_block_unblock(m, status, msg.message_id))


def start_admin_message_flow(call, back_kb):
    msg = bot.send_message(call.message.chat.id, "Код заказа или ID для связи:", reply_markup=back_kb)
    bot.register_next_step_handler(msg, process_msg_find, msg.message_id)


def handle_msg_client_callback(call, back_kb):
    code = call.data.split("|")[1]
    orders = load_data(ORDERS_FILE)
    if code not in orders:
        bot.send_message(call.message.chat.id, "❌ Заказ не найден.")
        return

    target_id = orders[code]['user_id']
    msg = bot.send_message(
        call.message.chat.id,
        f"Введите текст сообщения для клиента (заказ {code}):",
        reply_markup=back_kb
    )
    bot.register_next_step_handler(msg, lambda m: process_msg_send(m, target_id, msg.message_id))


def handle_client_info_callback(call):
    code = call.data.split("|")[1]
    orders = load_data(ORDERS_FILE)
    if code not in orders:
        bot.send_message(call.message.chat.id, "❌ Заказ не найден.")
        return

    target_id = orders[code]['user_id']
    user_orders = {current_code: data for current_code, data in orders.items() if data.get('user_id') == target_id}
    sorted_codes = get_sorted_order_codes(user_orders)

    lines = []
    for current_code in sorted_codes:
        order_data = user_orders[current_code]
        lines.append(
            f"🔹 <code>{current_code}</code> — {order_data['fio']} | {order_data['phone']} | {get_order_status_text(order_data)}"
        )

    text = f"👤 <b>Клиент <code>{target_id}</code>:</b>\n\n" + "\n".join(lines)
    bot.send_message(call.message.chat.id, text, parse_mode="HTML")


def handle_order_confirmation_callback(call):
    user_id = call.message.chat.id

    if call.data == "order_cancel":
        if user_id in temp_orders:
            del temp_orders[user_id]
        log_callback_event(call, "Пользователь отменил оформление заказа")
        replace_message_content(
            user_id,
            call.message.message_id,
            "❌ Оформление заказа отменено. Вы можете начать заново через меню."
        )
        return

    if user_id not in temp_orders:
        bot.send_message(user_id, "Время ожидания истекло или заказ уже обработан. Начните заново.")
        return

    data = temp_orders.pop(user_id)
    if not data.get("file_id") or order_has_invalid_user_fields(data):
        retry_text = (
            "❌ При оформлении заказа произошла ошибка. "
            "Пожалуйста, попробуйте сделать заказ снова через меню."
        )
        if not replace_message_content(user_id, call.message.message_id, retry_text):
            bot.send_message(user_id, retry_text)
        return

    rand_code = get_unique_code()
    orders = load_data(ORDERS_FILE)
    orders[rand_code] = make_order_record(user_id, rand_code, data)
    orders[rand_code]["admin_review_messages"] = send_order_for_admin_review(rand_code, orders[rand_code])
    save_data(ORDERS_FILE, orders)
    append_confirmed_order_to_daily_sheet(rand_code, orders[rand_code], "авто")
    log_callback_event(call, f"Пользователь подтвердил оформление заказа {rand_code}")

    replace_message_content(
        user_id,
        call.message.message_id,
        f"✅ Заказ оформлен! Ваш код получения: <code>{rand_code}</code>.\n"
        f"Он уже передан администраторам и добавлен в очередь заказов. "
        f"Когда заказ заберут и он приедет в ПВЗ, мы вам напишем."
    )


# --- CALLBACK LOGIC ---
@bot.callback_query_handler(func=lambda call: True)
def callback_logic(call):
    bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)
    back_kb = build_back_keyboard()
    log_callback_event(call, f"chat={call.message.chat.id}")

    if call.data == "admin_menu":
        open_admin_menu_from_callback(call)

    elif call.data.startswith("accepted_"):
        handle_taken_orders_action(call)

    elif call.data == "admin_receive":
        start_receive_flow(call, back_kb)

    elif call.data == "admin_find_by_id":
        start_find_client_flow(call, back_kb)

    elif call.data == "cancel_admin_msg":
        cancel_contact_admin_flow(call)

    elif call.data == "admin_giveout_start":
        start_giveout_flow(call, back_kb)

    elif call.data.startswith("admin_approve|") or call.data.startswith("admin_reject|"):
        handle_admin_review_action(call)

    elif call.data.startswith("give_confirm|"):
        handle_issue_confirmation(call, back_kb)

    elif call.data.startswith("orders_list|"):
        show_orders_list_page(call, int(call.data.split("|")[1]))

    elif call.data == "admin_info_search":
        start_order_info_search(call, back_kb)

    elif call.data.startswith("del_confirm|"):
        code = call.data.split("|")[1]
        orders = load_data(ORDERS_FILE)
        if code in orders:
            deleted_order = orders[code]
            del orders[code]
            save_data(ORDERS_FILE, orders)
            delete_message_safe(call.message.chat.id, call.message.message_id)
            bot.send_message(
                call.message.chat.id,
                f"✅ Заказ <code>{code}</code> удален.",
                parse_mode="HTML",
                reply_markup=back_kb
            )
            delete_order_and_notify_user(code, deleted_order)

    elif call.data in {"order_confirm", "order_cancel"}:
        handle_order_confirmation_callback(call)

    elif call.data == "admin_access":
        show_access_menu(call)

    elif call.data == "block_user":
        start_block_unblock_flow(call, True, back_kb)

    elif call.data == "unblock_user":
        start_block_unblock_flow(call, False, back_kb)

    elif call.data == "admin_msg":
        start_admin_message_flow(call, back_kb)

    elif call.data.startswith("msg_client|"):
        handle_msg_client_callback(call, back_kb)

    elif call.data.startswith("client_info|"):
        handle_client_info_callback(call)

# --- ЛОГИКА ШАГОВ (ПРИЕМКА И ВЫДАЧА) ---

def process_receive_mass(message, prompt_message_id=None):
    code = message.text
    orders = load_data(ORDERS_FILE)
    back_kb = build_back_keyboard()
    log_message_event(message, "receive_order_code_entered", f"code={sanitize_log_text(code)}")
    cleanup_transient_input(message, prompt_message_id)

    if code in orders:
        if orders[code].get('shelf'):
            prompt = bot.send_message(
                message.chat.id,
                f"⚠️ Этот заказ уже на полке № {orders[code]['shelf']}. Введите другой:",
                reply_markup=back_kb
            )
            bot.register_next_step_handler(prompt, process_receive_mass, prompt.message_id)
        elif has_order_price(orders[code]):
            log_message_event(message, "receive_order_using_saved_price", f"code={code}; price={orders[code].get('price')}")
            place_order_on_shelf(message.chat.id, code, orders, back_kb)
        else:
            msg = bot.send_message(
                message.chat.id,
                f"Введите цену для заказа <code>{code}</code> (числом):",
                parse_mode="HTML",
                reply_markup=back_kb
            )
            bot.register_next_step_handler(msg, lambda m: process_receive_price(m, code, msg.message_id))
    else:
        prompt = bot.send_message(
            message.chat.id,
            f"❌ Код <code>{code}</code> не найден. Проверьте и введите снова:",
            reply_markup=back_kb,
            parse_mode="HTML"
        )
        bot.register_next_step_handler(prompt, process_receive_mass, prompt.message_id)

def process_receive_price(message, code, prompt_message_id=None):
    back_kb = build_back_keyboard()
    price_str = message.text.replace(',', '.')
    log_message_event(message, "receive_order_price_entered", f"code={code}; price={sanitize_log_text(message.text)}")
    cleanup_transient_input(message, prompt_message_id)
    try:
        price = float(price_str)
    except ValueError:
        msg = bot.send_message(
            message.chat.id,
            "❌ Цена должна быть числом. Введите цену еще раз:",
            reply_markup=back_kb
        )
        bot.register_next_step_handler(msg, lambda m: process_receive_price(m, code, msg.message_id))
        return

    orders = load_data(ORDERS_FILE)
    if code not in orders:
        bot.send_message(message.chat.id, "❌ Заказ не найден.", reply_markup=back_kb)
        return

    orders[code]["price"] = price
    log_message_event(message, "order_received_to_pvz", f"code={code}; price={price}")
    place_order_on_shelf(message.chat.id, code, orders, back_kb)


def process_find_by_id(message, prompt_message_id=None):
    target_id_str = message.text.strip()
    log_message_event(message, "find_client_by_id", f"id={sanitize_log_text(target_id_str)}")
    cleanup_transient_input(message, prompt_message_id)
    back_to_menu_kb = build_back_keyboard("admin_menu", BTN_BACK_TO_MENU)
    if not target_id_str.isdigit():
        bot.send_message(message.chat.id, "❌ Ошибка: ID должен состоять только из цифр.", reply_markup=back_to_menu_kb)
        return

    target_id = int(target_id_str)
    orders = load_data(ORDERS_FILE)

    # ИСПОЛЬЗУЕМ НОВУЮ СОРТИРОВКУ С ФИЛЬТРОМ ПО ID
    sorted_user_codes = get_sorted_order_codes(orders, filter_ids=target_id)

    found_orders = []
    for code in sorted_user_codes:
        data = orders[code]
        shelf_info = f" ({get_order_status_text(data)})"
        found_orders.append(f"🔹 <code>{code}</code> — {data['fio']} | {data['phone']}{shelf_info}")
    # ... (остальной код функции без изменений)

    if found_orders:
        response = f"👤 <b>Заказы пользователя <code>{target_id}</code>:</b>\n\n" + "\n".join(found_orders)
        bot.send_message(message.chat.id, response, parse_mode="HTML", reply_markup=back_to_menu_kb)
    else:
        bot.send_message(message.chat.id, f"🔍 У пользователя с ID <code>{target_id}</code> заказов не найдено.",
                         parse_mode="HTML", reply_markup=back_to_menu_kb)

def process_giveout_search(message, prompt_message_id=None):
    code = message.text
    orders = load_data(ORDERS_FILE)
    back_kb = build_back_keyboard()
    log_message_event(message, "giveout_order_code_entered", f"code={sanitize_log_text(code)}")
    cleanup_transient_input(message, prompt_message_id)

    if code in orders:
        data = orders[code]
        if data.get('status') == STATUS_ISSUED:
            bot.send_message(message.chat.id,
                             f"⚠️ Заказ <code>{code}</code> уже выдан.",
                             parse_mode="HTML", reply_markup=back_kb)
            return

        if not data.get('shelf'):
            bot.send_message(message.chat.id,
                             f"⚠️ Заказ <code>{code}</code> еще не размещен на полке и не готов к выдаче.",
                             parse_mode="HTML", reply_markup=back_kb)
            return

        shelf_num = data.get('shelf')

        codes_on_shelf = get_order_codes_for_action(orders, code)

        total_price = 0
        has_missing_price = False
        for c in codes_on_shelf:
            p = orders[c].get('price')
            if p is not None:
                total_price += get_order_price_value(orders[c])
            else:
                has_missing_price = True

        total_commission = int(total_price // 10)
        price_text = f"💰 <b>К оплате (10%):</b> {total_commission} руб."
        if has_missing_price:
            price_text += " ⚠️ (у части заказов цена не указана)"

        codes_list = "\n".join(
            [f"  • <code>{c}</code> — {orders[c]['fio']} | {orders[c]['count']} шт." for c in codes_on_shelf])
        issue_title = (
            f"📦 <b>Будет выдан ({len(codes_on_shelf)}):</b>"
            if len(codes_on_shelf) == 1 else
            f"📦 <b>Будут выданы ({len(codes_on_shelf)}):</b>"
        )

        info = (f"📤 <b>ВЫДАЧА ЗАКАЗА</b>\n\n"
                f"🚪 <b>Ячейка: {shelf_num}</b>\n\n"
                f"{issue_title}\n{codes_list}\n\n"
                f"{price_text}")

        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("✅ Подтвердить выдачу", callback_data=f"give_confirm|{code}"))
        kb.add(types.InlineKeyboardButton("Отмена", callback_data="admin_menu"))

        bot.send_message(message.chat.id, info, parse_mode="HTML", reply_markup=kb)
    else:
        bot.send_message(message.chat.id, "❌ Код не найден.", reply_markup=back_kb)


# --- ОСТАЛЬНЫЕ ХЕНДЛЕРЫ ---

def process_info_search(message, prompt_message_id=None):
    code = message.text
    orders = load_data(ORDERS_FILE)
    log_message_event(message, "order_info_requested", f"code={sanitize_log_text(code)}")
    cleanup_transient_input(message, prompt_message_id)
    back_to_menu_kb = build_back_keyboard("admin_menu", BTN_BACK_TO_MENU)
    if code in orders:
        data = orders[code]
        price_value = data.get('price')
        price_info = f"💰 Цена: {price_value} руб." if price_value else "💰 Цена: не указана"
        info = build_order_detail_caption(
            "📋 <b>Информация о заказе</b>",
            code,
            data,
            status_text=get_order_status_info_text(data),
            extra_lines=[price_info]
        )
        bot.send_photo(
            message.chat.id,
            data['file_id'],
            caption=info,
            parse_mode="HTML",
            reply_markup=build_admin_order_info_keyboard(code)
        )
    else:
        bot.send_message(message.chat.id, "Не найден.", reply_markup=back_to_menu_kb)


def process_block_unblock(message, status, prompt_message_id=None):
    uid = message.text
    action_name = "block_user" if status else "unblock_user"
    log_message_event(message, action_name, f"target_id={sanitize_log_text(uid)}")
    cleanup_transient_input(message, prompt_message_id)
    users = load_data(USERS_FILE)
    if uid not in users: users[uid] = {}
    users[uid]['blocked'] = status
    save_data(USERS_FILE, users)
    bot.send_message(message.chat.id, "Готово.", reply_markup=build_admin_panel_keyboard())


def process_msg_find(message, prompt_message_id=None):
    input_val = message.text.strip()
    orders = load_data(ORDERS_FILE)
    target_id = None
    log_message_event(message, "admin_message_target_entered", f"value={sanitize_log_text(input_val)}")
    cleanup_transient_input(message, prompt_message_id)

    # Если 4 цифры - ищем код заказа
    if len(input_val) == 4 and input_val.isdigit():
        if input_val in orders:
            target_id = orders[input_val]['user_id']
            context = f"по заказу {input_val}"
        else:
            bot.send_message(message.chat.id, "❌ Код заказа не найден.")
            return admin_panel(message)
    # Если больше цифр - считаем за Telegram ID
    elif input_val.isdigit():
        target_id = int(input_val)
        context = f"по ID {input_val}"
    else:
        bot.send_message(message.chat.id, "❌ Неверный формат. Введите 4 цифры или ID.")
        return admin_panel(message)

    back_kb = build_back_keyboard()
    msg = bot.send_message(message.chat.id, f"Введите текст сообщения для клиента ({context}):", reply_markup=back_kb)
    bot.register_next_step_handler(msg, lambda m: process_msg_send(m, target_id, msg.message_id))


def process_taken_order_price_input(message, order_code, index, prompt_message_id=None, source_message_id=None):
    cleanup_transient_input(message, prompt_message_id)
    if is_menu_or_command(message.text):
        if message.text == "/admin":
            return admin_panel(message)
        return handle_text(message)

    price_str = message.text.replace(",", ".").strip()
    try:
        price = float(price_str)
    except ValueError:
        retry_prompt = bot.send_message(
            message.chat.id,
            f"❌ Стоимость для заказа <code>{order_code}</code> должна быть числом. Введите еще раз:",
            parse_mode="HTML"
        )
        bot.register_next_step_handler(
            retry_prompt,
            lambda m: process_taken_order_price_input(
                m,
                order_code,
                index,
                retry_prompt.message_id,
                source_message_id
            )
        )
        return

    orders = load_data(ORDERS_FILE)
    if order_code not in orders or orders[order_code].get("status") != STATUS_ACCEPTED_BY_ADMIN:
        if source_message_id is not None:
            show_taken_orders_view(message.chat.id, message.chat.id, index=index, message_id=source_message_id)
        return

    order_data = orders[order_code]
    order_data["price"] = price
    mark_order_as_in_transit(order_data, message.chat.id)
    save_data(ORDERS_FILE, orders)
    log_message_event(message, "today_order_accepted", f"code={order_code}; price={price}")

    try:
        bot.send_message(
            order_data["user_id"],
            f"🚚 Ваш заказ <code>{order_code}</code> забрали, он уже в пути в ПВЗ.",
            parse_mode="HTML"
        )
    except:
        pass

    if source_message_id is not None:
        show_taken_orders_view(message.chat.id, message.chat.id, index=index, message_id=source_message_id)


def process_taken_order_message_input(message, order_code, index, prompt_message_id=None):
    cleanup_prompt_messages(message.chat.id, prompt_message_id)
    if is_menu_or_command(message.text):
        if message.text == "/admin":
            return admin_panel(message)
        return handle_text(message)

    orders = load_data(ORDERS_FILE)
    if order_code not in orders:
        bot.send_message(message.chat.id, "❌ Заказ уже удален.")
        show_taken_orders_view(message.chat.id, message.chat.id, index=index)
        return

    log_message_event(message, "taken_order_custom_message_sent", f"code={order_code}; text={sanitize_log_text(message.text)}")
    target_id = orders[order_code]["user_id"]
    try:
        bot.send_message(target_id, f"Сообщение от админа по заказу <code>{order_code}</code>:\n{message.text}", parse_mode="HTML")
        bot.send_message(message.chat.id, "✅ Сообщение отправлено пользователю.")
    except:
        bot.send_message(message.chat.id, "❌ Не удалось отправить сообщение пользователю.")


def process_msg_send(message, user_id, prompt_message_id=None):
    cleanup_prompt_messages(message.chat.id, prompt_message_id)
    if is_menu_or_command(message.text):
        if message.text == "/admin":
            return admin_panel(message)
        return handle_text(message)

    log_message_event(message, "admin_message_sent_to_user", f"target_id={user_id}; text={sanitize_log_text(message.text)}")
    try:
        bot.send_message(user_id, f"Сообщение от админа: {message.text}")
        bot.send_message(message.chat.id, "✅ Отправлено.", reply_markup=build_admin_panel_keyboard())
    except:
        bot.send_message(message.chat.id, "❌ Не удалось отправить сообщение пользователю.", reply_markup=build_admin_panel_keyboard())

    sender_id = message.chat.id
    notify_text = (f"📨 <b>Админ <code>{sender_id}</code> написал клиенту <code>{user_id}</code>:</b>\n\n"
                   f"💬 {message.text}")
    for admin_id in ADMIN_IDS:
        if admin_id != sender_id:
            try:
                bot.send_message(admin_id, notify_text, parse_mode="HTML")
            except:
                pass

@bot.message_handler(func=lambda m: True)
def handle_text(message):
    if is_blocked(message.chat.id): return
    if message.text == BTN_MY_ORDERS:
        log_message_event(message, "open_my_orders")
        delete_message_safe(message.chat.id, message.message_id)
        orders = load_data(ORDERS_FILE)
        bot.send_message(message.chat.id, build_user_orders_text(message.chat.id, orders), parse_mode="HTML")
    elif message.text == BTN_CONTACT_ADMIN:
        log_message_event(message, "open_contact_admin")
        delete_message_safe(message.chat.id, message.message_id)
        kb = types.InlineKeyboardMarkup()
        kb.add(types.InlineKeyboardButton("⬅️ Назад", callback_data="cancel_admin_msg"))
        msg = bot.send_message(message.chat.id, "Введите текст вашего обращения:", reply_markup=kb)
        bot.register_next_step_handler(msg, forward_to_admins, msg.message_id)
    elif message.text == BTN_ADMIN_PANEL:
        log_message_event(message, "open_admin_panel", "Кнопка админ-панели")
        delete_message_safe(message.chat.id, message.message_id)
        if is_admin(message.chat.id):
            admin_panel(message)
    elif message.text == BTN_CREATE_ORDER:
        log_message_event(message, "start_create_order")
        delete_message_safe(message.chat.id, message.message_id)
        current_hour = datetime.now().hour
        if (current_hour >= 16 or datetime.now().weekday() == 6) and not is_admin(message.chat.id):
            bot.send_message(message.chat.id,
                             "К сожалению, пункт выдачи заказов сейчас не принимает заказы (принимает с 0:10 до 16:00, воскресенье выходной).\n"
                             "Пожалуйста сделайте заказ в рабочее время.\n\n"
                             "❗️ ВАЖНО: когда будете оформлять заказ в следующий раз, обновите QR-код в приложении, чтобы он был сегодняшним!")
            return
        bot.send_message(message.chat.id, "Пришлите фото QR-кода для оформления заказа:")
    elif message.text == "/start":
        log_message_event(message, "start_command")
        delete_message_safe(message.chat.id, message.message_id)
        main_menu(message)
    else:
        log_message_event(message, "plain_text_message", f"text={sanitize_log_text(message.text)}")


def forward_to_admins(message, prompt_message_id=None):
    cleanup_prompt_messages(message.chat.id, prompt_message_id)
    if is_menu_or_command(message.text):
        if message.text == "/admin":
            return admin_panel(message)
        return handle_text(message)

    log_message_event(message, "message_forwarded_to_admins", f"text={sanitize_log_text(message.text)}")
    notify_admins(f"📩 От <code>{message.chat.id}</code>:\n{message.text}")
    bot.send_message(
        message.chat.id,
        "Ваше сообщение отправлено администраторам.",
        reply_markup=build_main_menu_keyboard(message.chat.id)
    )

if __name__ == '__main__':
    print("Бот запущен...")
    bot.infinity_polling()
